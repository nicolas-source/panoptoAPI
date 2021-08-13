from openpyxl import Workbook
from openpyxl import load_workbook

workbook = load_workbook(filename='Automated Scheduling Test Local.xlsx', data_only=True)

# print(workbook['Panopto Folders to Create']['G2'].value)
ws = workbook['Panopto Folders to Create']
r = 1
# for row in ws.iter_rows('G{}:G{}'.format(ws.min_row, ws.max_row)):
#     for cell in row:
#         print(cell.value)
#         r += 1

for row in ws.iter_rows(min_row=1, min_col=7, max_row=ws.max_row, max_col=7):
    for cell in row:
        print(cell.value, end=" ")
    print()